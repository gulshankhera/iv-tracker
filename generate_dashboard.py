"""
Generate Excel Dashboard for IV Tracking
Creates a formatted Excel file with color-coded alerts
"""

import pandas as pd
import sqlite3
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime, timedelta
import os

DB_PATH = 'iv_data.db'


def create_dashboard(output_path='iv_dashboard.xlsx'):
    """Generate Excel dashboard with IV alerts and history"""

    conn = sqlite3.connect(DB_PATH)

    # Get recent alerts (last 30 days)
    cutoff = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')

    alerts_query = '''
        SELECT 
            ticker,
            alert_date as "Date",
            iv_change_pct as "IV Change %",
            iv_current as "Current IV",
            iv_previous as "Previous IV",
            analysis as "GPT Analysis"
        FROM alerts
        WHERE alert_date >= ?
        ORDER BY alert_date DESC, abs(iv_change_pct) DESC
    '''

    alerts_df = pd.read_sql_query(alerts_query, conn, params=(cutoff,))

    # Get latest IV for all tickers
    latest_query = '''
        WITH RankedIV AS (
            SELECT 
                ticker,
                date,
                iv_30day,
                price,
                ROW_NUMBER() OVER (PARTITION BY ticker ORDER BY date DESC) as rn
            FROM iv_history
        )
        SELECT 
            ticker as "Ticker",
            date as "Last Updated",
            iv_30day as "Current IV",
            price as "Price"
        FROM RankedIV
        WHERE rn = 1
        ORDER BY ticker
    '''

    current_df = pd.read_sql_query(latest_query, conn)

    # Calculate 7-day changes for current view
    changes = []
    for ticker in current_df['Ticker'].unique():
        change_query = '''
            SELECT iv_30day, date
            FROM iv_history
            WHERE ticker = ?
            ORDER BY date DESC
            LIMIT 8
        '''
        hist_df = pd.read_sql_query(change_query, conn, params=(ticker,))

        if len(hist_df) >= 2:
            change_pct = ((hist_df.iloc[0]['iv_30day'] - hist_df.iloc[-1]['iv_30day']) /
                          hist_df.iloc[-1]['iv_30day']) * 100
            changes.append(round(change_pct, 2))
        else:
            changes.append(0)

    current_df['7-Day Change %'] = changes

    conn.close()

    # Create Excel with pandas
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        current_df.to_excel(writer, sheet_name='Current IV', index=False)
        alerts_df.to_excel(writer, sheet_name='Alerts', index=False)

    # Apply formatting
    wb = load_workbook(output_path)
    format_current_sheet(wb['Current IV'])
    format_alerts_sheet(wb['Alerts'])

    wb.save(output_path)
    print(f"✅ Dashboard created: {output_path}")

    return output_path


def format_current_sheet(sheet):
    """Format the Current IV sheet with color coding"""

    # Header formatting
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Column widths
    sheet.column_dimensions['A'].width = 12  # Ticker
    sheet.column_dimensions['B'].width = 15  # Last Updated
    sheet.column_dimensions['C'].width = 12  # Current IV
    sheet.column_dimensions['D'].width = 12  # Price
    sheet.column_dimensions['E'].width = 15  # 7-Day Change

    # Data formatting
    for row in range(2, sheet.max_row + 1):
        # Ticker - bold
        sheet[f'A{row}'].font = Font(bold=True, size=10)
        sheet[f'A{row}'].alignment = Alignment(horizontal='center')

        # IV and Price formatting
        if sheet[f'C{row}'].value:
            sheet[f'C{row}'].number_format = '0.0"%"'
        if sheet[f'D{row}'].value:
            sheet[f'D{row}'].number_format = '$#,##0.00'

        # Change % with color coding
        change_cell = sheet[f'E{row}']
        if change_cell.value is not None:
            change_val = change_cell.value
            change_cell.number_format = '0.0"%"'

            # Red for IV spike (>20%), Green for crush (<-20%), Yellow for moderate
            if change_val > 20:
                change_cell.fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
                change_cell.font = Font(bold=True, color='FFFFFF')
            elif change_val < -20:
                change_cell.fill = PatternFill(start_color='51CF66', end_color='51CF66', fill_type='solid')
                change_cell.font = Font(bold=True, color='FFFFFF')
            elif abs(change_val) > 10:
                change_cell.fill = PatternFill(start_color='FFD93D', end_color='FFD93D', fill_type='solid')
                change_cell.font = Font(bold=True)

    # Freeze panes
    sheet.freeze_panes = 'A2'

    # Add borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row,
                               min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.border = thin_border


def format_alerts_sheet(sheet):
    """Format the Alerts sheet"""

    # Header formatting
    header_fill = PatternFill(start_color='E74C3C', end_color='E74C3C', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Column widths
    sheet.column_dimensions['A'].width = 12  # Ticker
    sheet.column_dimensions['B'].width = 15  # Date
    sheet.column_dimensions['C'].width = 15  # IV Change %
    sheet.column_dimensions['D'].width = 12  # Current IV
    sheet.column_dimensions['E'].width = 12  # Previous IV
    sheet.column_dimensions['F'].width = 60  # Analysis

    # Data formatting
    for row in range(2, sheet.max_row + 1):
        # Ticker
        sheet[f'A{row}'].font = Font(bold=True, size=10)
        sheet[f'A{row}'].alignment = Alignment(horizontal='center')

        # Change % with conditional formatting
        change_cell = sheet[f'C{row}']
        if change_cell.value is not None:
            change_val = change_cell.value
            change_cell.number_format = '0.0"%"'

            if change_val > 0:
                change_cell.fill = PatternFill(start_color='FFE5E5', end_color='FFE5E5', fill_type='solid')
                change_cell.font = Font(bold=True, color='C92A2A')
            else:
                change_cell.fill = PatternFill(start_color='E5FFE5', end_color='E5FFE5', fill_type='solid')
                change_cell.font = Font(bold=True, color='2B8A3E')

        # IV formatting
        for col in ['D', 'E']:
            if sheet[f'{col}{row}'].value:
                sheet[f'{col}{row}'].number_format = '0.0"%"'

        # Analysis text wrap
        sheet[f'F{row}'].alignment = Alignment(wrap_text=True, vertical='top')

    # Freeze panes
    sheet.freeze_panes = 'A2'

    # Borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row,
                               min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.border = thin_border


if __name__ == "__main__":
    if not os.path.exists(DB_PATH):
        print(f"❌ Database not found: {DB_PATH}")
        print("   Run iv_tracker.py first to collect data")
    else:
        output = create_dashboard()
        print(f"\n📊 Open {output} to view your IV dashboard")